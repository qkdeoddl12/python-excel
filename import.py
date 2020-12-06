from openpyxl import load_workbook

load_wb=load_workbook("과일.xlsx",data_only=True)

load_ws=load_wb['Sheet1']

print(load_ws['A1'].value)

print(load_ws.cell(1,2).value)

print('\n------지정한 셀 출력-----')
get_cells=load_ws['A1':'D2']
for row in get_cells:
    for cell in row:
        print(cell.value)


print('\n------모든 행 단위로 출력-----')

for row in load_ws.rows:
        print(row)